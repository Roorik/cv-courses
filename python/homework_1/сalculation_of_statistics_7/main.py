# У вас есть Excel-файл `data.xlsx` с данными о продажах за месяц. 
# Необходимо вычислить общую сумму продаж и среднюю продажу за день.
import logging

from pathlib import Path
from openpyxl import load_workbook, Workbook

# определяем директорию под чтение и запись файлов, логов
BASE_DIR = Path(__file__).parent
input_dir = BASE_DIR.joinpath('data.xlsx')
log_dir = BASE_DIR.joinpath('processing_log.txt')
output_dir = BASE_DIR.joinpath('sales_report.xlsx')

# настраиваем логирование
logging.basicConfig(filename=log_dir, level=logging.INFO, filemode='w')
log = logging.getLogger(__name__)

# здесь мы считаем общую и среднюю сумму 
def calculate_sales(data):
    total = 0 # общая сумма продаж 
    avarage = 0 # средняя сумма 
    units = 0 # количество продаж
    
    # находим индекс суммы продажи
    sales_index = None   
    for cell in data[1]:
        # поле, которое нам необходимо посчитать
        if cell.value == 'Общая сумма продажи':
            sales_index = cell.column - 1
            break
        
    # если нужное нам поле есть 
    if sales_index is not None:
        # перебираем строки листа 
        for row in data.iter_rows(min_row = 2, values_only = True):
            # инкрементируем количество продаж
            units += 1
            # считаем общую сумму
            total += row[sales_index]
    else:
        log.error(f'Поле "Общая сумму продажи" в файле {input_dir} отсутствует')
    # считаем среднюю сумму
    avarage = total/units
    
    # логируем успешный результат
    log.info(f'Результат вычислен успешно, общая сумма: {total}, средняя за день: {avarage}')

    return total, avarage


if __name__ == '__main__':
    # выполняем программу
    try:
        # пытаемся открыть файл по заданной директории с входными данными
        try:
            # выгружаем данные в переменную 
            workbook = load_workbook(input_dir)
            data = workbook.active
            log.info(f'Файл {input_dir} успешно открыт')
        except:
            log.error(f'Файл {input_dir} не найден')
        total_sum, avarage_sum = calculate_sales(data)
        
        # создаём новый excel файл
        new_workbook = Workbook()
        new_sheet = new_workbook.active
        new_sheet.title = "Итоги"
        
        # запись результата
        new_sheet['A1'] = 'Общая сумма продаж за месяц:'
        new_sheet['B1'] = round(total_sum, 2)
        
        new_sheet['A2'] = 'Средняя сумма продаж за день:'
        new_sheet['B2'] = round(avarage_sum, 2)
        
        # сохранение файла
        new_workbook.save(output_dir)        
        log.info(f'Программа отработала без ошибок, результат записан в {output_dir}')
        
    # если в программе что-то пойдёт не так
    except Exception as e:
        log.error(f'Произошла ошибка: {e}')